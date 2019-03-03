# -*- coding: utf-8 -*-
import re
from datetime import datetime
from typing import Tuple, List, Dict, Any

import pandas
import xlrd
from openpyxl import load_workbook


def load_wb_data(
    path,
    sheet_index: int = 0,
    table_row: int = 0,
    table_col: int = 0,
    invert_direction: bool = False,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = xlrd.open_workbook(path, on_demand=True)
    ws = wb.sheet_by_index(sheet_index)
    header = []
    data = []

    cols = ws.ncols
    rows = ws.nrows
    table_range = table_row + 1, rows

    if invert_direction:
        cols = ws.nrows
        rows = ws.ncols
        table_range = table_col + 1, rows

    for col in range(cols):
        cell_value = ws.cell_value(table_row, col)

        if invert_direction:
            cell_value = ws.cell_value(col, table_row)

        cell_value = " ".join(str(cell_value).split()).lower()
        cell_value = re.sub(r"[^\w]", "_", cell_value).strip("_")
        header.append(cell_value)

    for row in range(*table_range):
        elm = {}
        for col in range(cols):
            cell_value = ws.cell_value(row, col)
            if invert_direction:
                cell_value = ws.cell_value(col, row)

            elm[header[col]] = cell_value

        data.append(elm)

    return header, data


def load_wb_cell(path, sheet_index=0, cell="A1"):
    wb = load_workbook(str(path))
    ws_name = wb.sheetnames[sheet_index]
    ws = wb[ws_name]

    return ws[cell].value


def convert_excel_time(excel_time):
    if isinstance(excel_time, float):
        return pandas.to_datetime("1899-12-30") + pandas.to_timedelta(excel_time, "D")

    if isinstance(excel_time, str):
        try:
            return datetime.strptime(excel_time, "%a, %d %b %Y %H:%M:%S %z")

        except ValueError:
            pass

    return None
